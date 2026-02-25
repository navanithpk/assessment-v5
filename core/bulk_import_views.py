"""
Bulk Import Views - For importing users from XLSX / XLS / CSV files
"""
import re
import datetime
import pandas as pd

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.db import transaction
from django.http import HttpResponse

from .models import UserProfile, Student, School, Grade, Subject


# ── helpers ──────────────────────────────────────────────────────────────────

def _safe_str(val):
    """Convert any cell value to a clean string (handles NaN, floats, ints)."""
    if val is None:
        return ''
    import math
    if isinstance(val, float) and math.isnan(val):
        return ''
    # pandas / numpy ints come as numpy.int64 etc – convert to native int first
    try:
        val = val.item()   # numpy scalar → python scalar
    except AttributeError:
        pass
    return str(val).strip()


def _next_roll(school, grade_obj, section):
    """Return the next sequential roll number string for a given class."""
    existing = Student.objects.filter(
        school=school, grade=grade_obj, section=section
    ).values_list('roll_number', flat=True)
    max_num = 0
    for roll in existing:
        if roll:
            m = re.search(r'(\d+)$', roll)
            if m:
                max_num = max(max_num, int(m.group(1)))
    return str(max_num + 1)


def _next_admission_id(school):
    """Return the next sequential ADM{year}### string for the school."""
    year = datetime.datetime.now().year
    prefix = f"ADM{year}"
    existing = Student.objects.filter(school=school).values_list('admission_id', flat=True)
    max_num = 0
    for adm in existing:
        if adm:
            if adm.startswith(prefix):
                try:
                    max_num = max(max_num, int(adm[len(prefix):]))
                except ValueError:
                    pass
            else:
                m = re.search(r'(\d+)$', adm)
                if m:
                    max_num = max(max_num, int(m.group(1)))
    return f"{prefix}{str(max_num + 1).zfill(3)}"


def _read_excel(file_obj):
    """Read an uploaded file (xlsx/xls/csv) into a DataFrame robustly."""
    name = getattr(file_obj, 'name', '').lower()
    if name.endswith('.csv'):
        df = pd.read_csv(file_obj, dtype=str)
    else:
        # dtype=str prevents pandas from auto-converting numbers → floats
        df = pd.read_excel(file_obj, dtype=str)
    # Normalise column names: strip whitespace, lowercase
    df.columns = [c.strip().lower().replace(' ', '_') for c in df.columns]
    return df


# ── views ─────────────────────────────────────────────────────────────────────

@login_required
def bulk_import_users(request):
    """Bulk import users from XLSX / XLS / CSV file."""
    if request.user.profile.role not in ['teacher', 'school_admin']:
        messages.error(request, "You don't have permission to import users.")
        return redirect('teacher_dashboard')

    school = request.user.profile.school

    if request.method == 'POST':

        # ── STEP 1: PREVIEW ────────────────────────────────────────────────
        if 'preview' in request.POST:
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "Please upload an Excel file.")
                return redirect('bulk_import_users')

            try:
                df = _read_excel(excel_file)
            except Exception as e:
                messages.error(request, f"Could not read file: {e}")
                return redirect('bulk_import_users')

            # Check required columns (case-insensitive, already lowercased)
            required_cols = ['username', 'email', 'password', 'first_name', 'last_name', 'role']
            missing = [c for c in required_cols if c not in df.columns]
            if missing:
                messages.error(
                    request,
                    f"Missing required columns: {', '.join(missing)}. "
                    f"Columns found: {', '.join(df.columns.tolist())}"
                )
                return redirect('bulk_import_users')

            # Fill NaN → '' and convert to records
            preview_data = df.fillna('').astype(str).to_dict('records')
            request.session['import_data'] = preview_data

            return render(request, 'teacher/bulk_import_preview.html', {
                'preview_data': preview_data[:50],
                'total_rows': len(preview_data),
            })

        # ── STEP 2: CONFIRM IMPORT ─────────────────────────────────────────
        elif 'confirm_import' in request.POST:
            import_data = request.session.get('import_data', [])
            if not import_data:
                messages.error(request, "No data to import. Please upload the file again.")
                return redirect('bulk_import_users')

            success_count = 0
            error_count = 0
            errors = []

            with transaction.atomic():
                for idx, row in enumerate(import_data, start=2):
                    try:
                        username   = _safe_str(row.get('username', ''))
                        email      = _safe_str(row.get('email', ''))
                        password   = _safe_str(row.get('password', ''))
                        first_name = _safe_str(row.get('first_name', ''))
                        last_name  = _safe_str(row.get('last_name', ''))
                        role       = _safe_str(row.get('role', '')).lower()

                        # ── Basic validation ──────────────────────────────
                        if not all([username, email, password, first_name, last_name, role]):
                            errors.append(f"Row {idx}: Missing required fields")
                            error_count += 1
                            continue

                        if role not in ['student', 'teacher']:
                            errors.append(
                                f"Row {idx}: Invalid role '{role}' — must be 'student' or 'teacher'"
                            )
                            error_count += 1
                            continue

                        if User.objects.filter(username=username).exists():
                            errors.append(f"Row {idx}: Username '{username}' already exists")
                            error_count += 1
                            continue

                        if User.objects.filter(email=email).exists():
                            errors.append(f"Row {idx}: Email '{email}' already exists")
                            error_count += 1
                            continue

                        # ── Create Django User ────────────────────────────
                        user = User.objects.create_user(
                            username=username,
                            email=email,
                            password=password,
                            first_name=first_name,
                            last_name=last_name,
                        )
                        if role == 'teacher':
                            user.is_staff = True
                            user.save()

                        # ── Resolve grade / section (students) ────────────
                        grade_name  = _safe_str(row.get('grade', ''))   or 'General'
                        section     = _safe_str(row.get('division', '')) or 'A'

                        # Attempt to extract numeric grade for UserProfile.grade
                        grade_num = None
                        try:
                            grade_num = int(grade_name)
                        except (ValueError, TypeError):
                            pass

                        # ── Create UserProfile ────────────────────────────
                        profile = UserProfile.objects.create(
                            user=user,
                            role=role,
                            school=school,
                            grade=grade_num,
                            division=section if role == 'student' else None,
                        )

                        # ── Student-specific record ───────────────────────
                        if role == 'student':
                            full_name    = _safe_str(row.get('full_name', ''))   or f"{first_name} {last_name}"
                            roll_number  = _safe_str(row.get('roll_number', ''))
                            admission_id = _safe_str(row.get('admission_id', ''))

                            grade_obj, _ = Grade.objects.get_or_create(name=grade_name)

                            # Auto-assign roll number if blank
                            if not roll_number:
                                roll_number = _next_roll(school, grade_obj, section)

                            # Auto-assign admission ID if blank
                            if not admission_id:
                                admission_id = _next_admission_id(school)

                            Student.objects.create(
                                user=user,
                                school=school,
                                full_name=full_name,
                                roll_number=roll_number,
                                admission_id=admission_id,
                                grade=grade_obj,
                                section=section,
                                created_by=request.user,
                            )

                        # ── Teacher subject ───────────────────────────────
                        if role == 'teacher':
                            subject = _safe_str(row.get('subject', ''))
                            if subject:
                                profile.subject = subject
                                profile.save()

                        success_count += 1

                    except Exception as e:
                        errors.append(f"Row {idx}: {e}")
                        error_count += 1

            # Clear session
            request.session.pop('import_data', None)

            if success_count:
                messages.success(request, f"✅ Successfully imported {success_count} user(s).")
            if error_count:
                messages.warning(request, f"⚠️ {error_count} row(s) had errors (shown below).")
                for err in errors[:10]:
                    messages.error(request, err)

            return redirect('manage_users')

    return render(request, 'teacher/bulk_import_users.html')


@login_required
def download_user_template(request):
    """Generate and download Excel template for bulk user import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # ── Students sheet ──
    ws_s = wb.active
    ws_s.title = "Students"

    student_headers = [
        'username', 'email', 'password', 'first_name', 'last_name',
        'role', 'full_name', 'roll_number', 'admission_id', 'grade', 'division'
    ]

    for col, hdr in enumerate(student_headers, 1):
        cell = ws_s.cell(row=1, column=col, value=hdr)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    sample_students = [
        ['alice.johnson', 'alice@school.com', 'Pass@1234', 'Alice', 'Johnson',
         'student', 'Alice Johnson', '', '', 'AS Level', 'A'],
        ['bob.williams',  'bob@school.com',   'Pass@1234', 'Bob',   'Williams',
         'student', 'Bob Williams',  '', '', 'AS Level', 'A'],
        ['charlie.brown', 'charlie@school.com','Pass@1234', 'Charlie','Brown',
         'student', 'Charlie Brown', '', '', 'A Level',  'B'],
    ]

    for r, row in enumerate(sample_students, 2):
        for c, val in enumerate(row, 1):
            ws_s.cell(row=r, column=c, value=val)

    # ── Teachers sheet ──
    ws_t = wb.create_sheet("Teachers")

    teacher_headers = [
        'username', 'email', 'password', 'first_name', 'last_name', 'role', 'subject'
    ]

    for col, hdr in enumerate(teacher_headers, 1):
        cell = ws_t.cell(row=1, column=col, value=hdr)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    sample_teachers = [
        ['john.physics', 'john.p@school.com', 'Teach@123', 'John', 'Peterson', 'teacher', 'Physics'],
        ['mary.chem',    'mary.c@school.com',  'Teach@123', 'Mary', 'Chen',     'teacher', 'Chemistry'],
    ]

    for r, row in enumerate(sample_teachers, 2):
        for c, val in enumerate(row, 1):
            ws_t.cell(row=r, column=c, value=val)

    # Auto-fit columns
    for ws in [ws_s, ws_t]:
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value is not None),
                default=10
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=bulk_user_import_template.xlsx'
    wb.save(response)
    return response
